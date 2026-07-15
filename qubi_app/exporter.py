from __future__ import annotations

import re
from copy import copy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from .constants import CATEGORY_CAPACITY, COST_PARAMETERS
from .validation import cleaned_rows

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "template" / "Fondo_QuBi_Scheda_Monitoraggio.xlsx"

ACTIVITY_ROWS = {
    "Alimentare": [18, 19, 20, 21, 22],
    "Scuola": [24, 25, 26, 27, 28, 29, 30, 31],
    "Sport": [33, 34, 35],
    "Offerta generale": [37, 38, 39, 40, 41, 42, 43, 44],
    "Azioni trasversali di coordinamento": [46, 47],
    "Ulteriori azioni progettuali": [49, 50, 51],
}
EXPENSE_ROWS = list(range(58, 62))
BENEFICIARY_ROWS = list(range(73, 79))


def _as_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return datetime.fromisoformat(str(value)).date()
    except ValueError:
        return None


def _as_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return int(number) if number.is_integer() else number


def _clear_range(ws, row: int, start_col: int, end_col: int) -> None:
    for column in range(start_col, end_col + 1):
        ws.cell(row=row, column=column).value = None


def _write_date(cell, value: Any) -> None:
    parsed = _as_date(value)
    cell.value = parsed
    if parsed:
        cell.number_format = "DD/MM/YYYY"


def _auto_cost_parameter(activity: str, current: Any) -> str:
    current_text = str(current or "").strip()
    if current_text:
        return current_text
    return COST_PARAMETERS.get(activity, "")


def build_excel(data: dict[str, Any], template_path: Path = TEMPLATE_PATH) -> bytes:
    workbook = load_workbook(template_path)
    ws = workbook["Scheda Monitoraggio"]

    general = data["general"]
    ws["B4"] = general.get("project_title", "")
    ws["B5"] = general.get("project_code", "")
    ws["B6"] = general.get("lead_org", "")
    _write_date(ws["B7"], general.get("project_start"))
    _write_date(ws["B8"], general.get("project_end"))
    ws["C9"] = str(general.get("contact_phone", "") or "")
    ws["C9"].number_format = "@"
    ws["C10"] = str(general.get("contact_email", "") or "")
    ws["C10"].number_format = "@"

    # Correzioni dei refusi presenti nel modello di origine.
    ws["D16"] = "periodo\n01/01/2026 - 30/06/2026"
    ws["N72"] = "Numero beneficiari raggiunti al 31/05/2027"
    ws["A46"] = "Coordinamento, comunicazione"

    for rows in ACTIVITY_ROWS.values():
        for row in rows:
            _clear_range(ws, row, 1, 7)

    grouped: dict[str, list[dict[str, Any]]] = {category: [] for category in ACTIVITY_ROWS}
    for row in cleaned_rows(data["activities_df"], "activity"):
        category = str(row.get("category") or "").strip()
        if category in grouped:
            grouped[category].append(row)

    for category, rows in grouped.items():
        for destination_row, record in zip(ACTIVITY_ROWS[category], rows):
            ws.cell(destination_row, 1).value = record.get("activity", "")
            _write_date(ws.cell(destination_row, 2), record.get("start_date"))
            for column, period in zip(range(3, 7), ("p1", "p2", "p3", "p4")):
                ws.cell(destination_row, column).value = record.get(f"{period}_status", "") or ""
            _write_date(ws.cell(destination_row, 7), record.get("end_date"))

    for row in EXPENSE_ROWS:
        _clear_range(ws, row, 1, 10)
    expenses = cleaned_rows(data["expenses_df"], "activity")
    for destination_row, record in zip(EXPENSE_ROWS, expenses):
        activity = str(record.get("activity") or "")
        ws.cell(destination_row, 1).value = activity
        ws.cell(destination_row, 2).value = _auto_cost_parameter(activity, record.get("cost_parameter"))
        columns = [
            (3, "p1_quantity"), (4, "p1_amount"),
            (5, "p2_quantity"), (6, "p2_amount"),
            (7, "p3_quantity"), (8, "p3_amount"),
            (9, "p4_quantity"), (10, "p4_amount"),
        ]
        for column, key in columns:
            ws.cell(destination_row, column).value = _as_number(record.get(key))
        for column in (4, 6, 8, 10):
            ws.cell(destination_row, column).number_format = '#,##0.00 [$€-it-IT]'

    contribution = data["contribution"]
    for column, period in zip(range(2, 6), ("p1", "p2", "p3", "p4")):
        ws.cell(66, column).value = float(contribution.get(period, 0) or 0) / 100
        ws.cell(66, column).number_format = "0%"

    for row in BENEFICIARY_ROWS:
        _clear_range(ws, row, 1, 16)
    beneficiaries = cleaned_rows(data["beneficiaries_df"], "activity")
    for destination_row, record in zip(BENEFICIARY_ROWS, beneficiaries):
        keys = [
            "activity", "beneficiary_type", "age_range", "target",
            "p1_reached", "p1_minors", "p1_relation",
            "p2_reached", "p2_minors", "p2_relation",
            "p3_reached", "p3_minors", "p3_relation",
            "p4_reached", "p4_minors", "p4_relation",
        ]
        numeric_keys = {
            "target", "p1_reached", "p1_minors", "p2_reached", "p2_minors",
            "p3_reached", "p3_minors", "p4_reached", "p4_minors",
        }
        for column, key in enumerate(keys, start=1):
            value = record.get(key)
            ws.cell(destination_row, column).value = _as_number(value) if key in numeric_keys else (value or "")
            ws.cell(destination_row, column).alignment = copy(ws.cell(73, column).alignment)

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A16"
    ws.print_area = "A1:P79"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def safe_filename(project_title: str, project_code: str) -> str:
    base = f"Scheda_Monitoraggio_QuBi_{project_code}_{project_title}".strip("_")
    base = re.sub(r"[^A-Za-z0-9À-ÿ_-]+", "_", base)
    base = re.sub(r"_+", "_", base).strip("_")
    return (base or "Scheda_Monitoraggio_QuBi") + ".xlsx"
